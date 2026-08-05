"""Extract structured text from a specification document.

The STD Generator agent reads plain text, so extraction turns a ``.docx``,
``.xlsx``, or ``.pdf`` specification into a single text representation that
preserves the document's structure (headings, paragraphs, tables, sheets,
pages). This is deterministic and dependency-only — no LLM involved.
"""

from __future__ import annotations

from pathlib import Path

import docx
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

SUPPORTED = (".docx", ".xlsx", ".pdf")


def extract(path: str | Path) -> str:
    """Return the text of a specification document.

    Raises ``FileNotFoundError`` if the file is missing and ``ValueError`` for
    an unsupported extension.
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"Specification file not found: {p}")

    suffix = p.suffix.lower()
    if suffix == ".docx":
        return _extract_docx(p)
    if suffix == ".xlsx":
        return _extract_xlsx(p)
    if suffix == ".pdf":
        return _extract_pdf(p)
    raise ValueError(
        f"Unsupported specification format: {suffix!r} (supported: {', '.join(SUPPORTED)})"
    )


def _extract_docx(path: Path) -> str:
    document = docx.Document(str(path))
    lines: list[str] = []
    for block in _iter_block_items(document):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if text:
                lines.append(text)
        else:  # Table
            for row in block.rows:
                cells = [cell.text.strip() for cell in row.cells]
                if any(cells):
                    lines.append(" | ".join(cells))
    return "\n".join(lines)


def _iter_block_items(document: docx.document.Document):
    """Yield paragraphs and tables in the order they appear in the body."""
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            yield Paragraph(child, document)
        elif child.tag.endswith("}tbl"):
            yield Table(child, document)


def _extract_xlsx(path: Path) -> str:
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        lines = [f"# Sheet: {sheet.title}"]
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                lines.append(" | ".join(cells))
        sections.append("\n".join(lines))
    workbook.close()
    return "\n\n".join(sections)


def _extract_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        pages.append(f"[Page {i}]\n{text}")
    return "\n\n".join(pages)
