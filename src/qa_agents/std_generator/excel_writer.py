"""Build the STD Excel workbook from an ``StdResult``.

The engine — not the LLM — writes the file, so the output is deterministic and
fully testable. Layout comes from a :class:`Profile` (defaults to CRM_HEBREW).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from ..models import StdResult
from .profile import CRM_HEBREW, ROW_NUMBER, Profile, SheetSpec


def write_workbook(
    result: StdResult, path: str | Path, profile: Profile = CRM_HEBREW
) -> Path:
    """Write ``result`` to an ``.xlsx`` file and return its path."""
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default empty sheet

    _write_sheet(workbook, profile.scenarios, result.scenarios, profile.rtl)
    _write_sheet(workbook, profile.sql, result.sql_queries, profile.rtl)

    out = Path(path)
    workbook.save(out)
    return out


def _write_sheet(
    workbook: Workbook, spec: SheetSpec, items: Sequence[Any], rtl: bool
) -> None:
    sheet: Worksheet = workbook.create_sheet(title=spec.title)
    sheet.sheet_view.rightToLeft = rtl
    align_header = Alignment(horizontal="right" if rtl else "left")
    align_cell = Alignment(horizontal="right" if rtl else "left", wrap_text=True)

    headers = [header for header, _ in spec.columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = align_header

    for row_number, item in enumerate(items, start=1):
        row = [_cell_value(source, item, row_number) for _, source in spec.columns]
        sheet.append(row)
        for cell in sheet[sheet.max_row]:
            cell.alignment = align_cell


def _cell_value(source: str | None, item: Any, row_number: int) -> Any:
    if source == ROW_NUMBER:
        return row_number
    if source is None:  # manual column, left blank
        return None
    return getattr(item, source)
