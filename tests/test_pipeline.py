"""End-to-end test: spec document → STD workbook, with a mocked LLM."""

import json

import docx
from openpyxl import load_workbook

from qa_agents.std_generator.agent import StdGeneratorAgent
from qa_agents.std_generator.pipeline import default_output_name, generate_std

_CANNED = json.dumps(
    {
        "scenarios": [
            {
                "entity": "הזמנה (demo_order)",
                "event": "עדכון סטטוס - חיובי",
                "target_field": 'שדה "סטטוס הזמנה" (demo_orderstatus)',
                "schema": "הזמנה (demo_order)",
                "condition": "demo_orderstatus = 1 וקיים demo_supplierid",
                "expected_result": "הסטטוס עובר ל-אושר (2)",
            }
        ],
        "sql_queries": [
            {
                "tag": "40012",
                "purpose": "שליפת אוכלוסייה לעדכון",
                "main_table": "demo_order",
                "sql": "SELECT * FROM demo_order WHERE demo_orderstatus = 1 AND statecode = 0",
                "notes": "",
            }
        ],
    },
    ensure_ascii=False,
)


def test_default_output_name_encodes_selection():
    assert default_output_name("1", True, True) == "output/STD_1_scenarios_sql.xlsx"
    assert default_output_name("1", True, False) == "output/STD_1_scenarios.xlsx"
    assert default_output_name("1", False, True) == "output/STD_1_sql.xlsx"


def test_generate_std_from_docx_produces_workbook(tmp_path):
    # a minimal synthetic spec
    document = docx.Document()
    document.add_paragraph("תיוג 40012: עדכון סטטוס הזמנה כאשר מתקבל אישור מספק.")
    spec = tmp_path / "spec.docx"
    document.save(spec)

    agent = StdGeneratorAgent(completer=lambda system, user: _CANNED)
    out = generate_std(spec, tag="40012", output_path=tmp_path / "std.xlsx", agent=agent)

    wb = load_workbook(out)
    assert wb.sheetnames == ["תסריטים", "SQL"]
    assert wb["תסריטים"].cell(row=2, column=2).value == "הזמנה (demo_order)"
    assert wb["SQL"].cell(row=2, column=5).value.startswith("SELECT")
