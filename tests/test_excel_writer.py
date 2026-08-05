"""Tests for the Excel engine — round-trips a workbook and inspects it."""

from openpyxl import load_workbook

from qa_agents.models import Scenario, SqlQuery, StdResult
from qa_agents.std_generator.excel_writer import write_workbook


def _sample_result() -> StdResult:
    return StdResult(
        scenarios=[
            Scenario(
                entity="פרטי (contact)",
                event="שליפת אוכלוסייה - חיובי",
                target_field='שדה "סירוב דוחות" (new_mailreports)',
                schema="פרטי (contact)",
                condition="new_mailreports = 0 או NULL",
                expected_result="רשומה תיכלל באוכלוסייה",
            ),
            Scenario(
                entity="פרטי (contact)",
                event="שליפת אוכלוסייה - שלילי",
                target_field='שדה "סירוב דוחות" (new_mailreports)',
                schema="פרטי (contact)",
                condition="new_mailreports = 1",
                expected_result="רשומה לא תיכלל",
            ),
        ],
        sql_queries=[
            SqlQuery(
                tag="12345",
                purpose="שליפת אוכלוסייה חיובית",
                main_table="contact",
                sql="SELECT * FROM contact WHERE statecode = 0",
                notes="",
            )
        ],
    )


def test_writes_two_named_rtl_sheets(tmp_path):
    out = write_workbook(_sample_result(), tmp_path / "std.xlsx")
    wb = load_workbook(out)

    assert wb.sheetnames == ["תסריטים", "SQL"]
    assert wb["תסריטים"].sheet_view.rightToLeft is True
    assert wb["SQL"].sheet_view.rightToLeft is True


def test_scenario_headers_and_values(tmp_path):
    out = write_workbook(_sample_result(), tmp_path / "std.xlsx")
    ws = load_workbook(out)["תסריטים"]

    headers = [c.value for c in ws[1]]
    assert headers[:7] == [
        "מס'", "ישות", "אירוע", "שדה", "סכמה", "תנאי/פעולה", "תוצאה צפויה",
    ]
    # first data row: auto number + agent-filled fields
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "פרטי (contact)"
    assert ws.cell(row=2, column=7).value == "רשומה תיכלל באוכלוסייה"
    # second scenario is numbered 2
    assert ws.cell(row=3, column=1).value == 2


def test_manual_columns_left_blank(tmp_path):
    out = write_workbook(_sample_result(), tmp_path / "std.xlsx")
    ws = load_workbook(out)["תסריטים"]

    # columns 8-11 (תוצאת בדיקה, הערות, ת.הרצה, סבב) are manual → empty
    for col in range(8, 12):
        assert ws.cell(row=2, column=col).value is None


def test_creates_missing_output_directory(tmp_path):
    out = write_workbook(_sample_result(), tmp_path / "output" / "STD_40012.xlsx")
    assert out.exists()


def test_sql_sheet_values(tmp_path):
    out = write_workbook(_sample_result(), tmp_path / "std.xlsx")
    ws = load_workbook(out)["SQL"]

    headers = [c.value for c in ws[1]]
    assert headers == ["מס'", "תיוג", "מטרת השאילתה", "טבלה ראשית", "שאילתת SQL", "הערות"]
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=5).value == "SELECT * FROM contact WHERE statecode = 0"
